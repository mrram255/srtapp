from __future__ import annotations

import io
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook

from apps.academic.models import CurriculumSubject, Program, Regulation


class SubjectBulkService:
    @staticmethod
    def import_template() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                'program_code',
                'code',
                'name',
                'semester_number',
                'credits',
                'subject_type',
                'category',
                'regulation_code',
            ]
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    @transaction.atomic
    def bulk_import(file_obj, college) -> dict[str, Any]:
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(h).strip().lower() for h in (header_row or []) if h is not None]
        required = {'program_code', 'code', 'name', 'semester_number'}
        if not required.issubset(set(headers)):
            raise ValueError(f'Missing columns. Required: {", ".join(sorted(required))}')

        created = 0
        errors: list[dict] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            data = dict(zip(headers, row))
            try:
                program = Program.objects.get(
                    college=college,
                    code=str(data['program_code']).strip(),
                    is_active=True,
                )
                regulation = None
                reg_code = data.get('regulation_code')
                if reg_code:
                    regulation = Regulation.objects.filter(
                        college=college,
                        code=str(reg_code).strip(),
                        is_active=True,
                    ).first()
                CurriculumSubject.objects.create(
                    college=college,
                    program=program,
                    regulation=regulation,
                    code=str(data['code']).strip(),
                    name=str(data['name']).strip(),
                    semester_number=int(data['semester_number']),
                    credits=int(data.get('credits') or 0),
                    subject_type=str(data.get('subject_type') or 'theory'),
                    category=str(data.get('category') or 'core'),
                )
                created += 1
            except Exception as exc:
                errors.append({'row': idx, 'error': str(exc)})
        return {'created': created, 'errors': errors}
