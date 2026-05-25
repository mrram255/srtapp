from __future__ import annotations

import io
import logging
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.core.utils import encrypt_field, generate_employee_id, validate_aadhaar
from apps.users.models import ModulePermission, Role, RolePermission, UserActivity

User = get_user_model()
logger = logging.getLogger(__name__)


class UserService:
    @staticmethod
    def check_permission(user, module: str, action: str) -> bool:
        if not user or not user.is_authenticated:
            return False
        if user.is_superuser or user.role in {'SUPER_ADMIN'}:
            return True
        if not user.role_ref_id:
            return False
        return user.role_ref.permissions.filter(
            module__name=module,
            action=action,
        ).exists()

    @staticmethod
    def get_accessible_modules(user) -> list[dict[str, Any]]:
        if user.is_superuser or user.role in {'SUPER_ADMIN'}:
            from apps.users.models import Module

            return list(
                Module.objects.filter(parent_module__isnull=True)
                .order_by('order')
                .values('name', 'display_name', 'icon', 'order')
            )
        if not user.role_ref_id:
            return []
        modules = (
            user.role_ref.permissions.select_related('module')
            .filter(action='view')
            .values('module__name', 'module__display_name', 'module__icon', 'module__order')
            .distinct()
        )
        return [
            {
                'name': row['module__name'],
                'display_name': row['module__display_name'],
                'icon': row['module__icon'],
                'order': row['module__order'],
            }
            for row in modules
        ]

    @staticmethod
    def log_activity(
        user,
        action: str,
        module: str = '',
        description: str = '',
        request=None,
        metadata: dict | None = None,
    ):
        ip_address = None
        user_agent = ''
        if request is not None:
            ip_address = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get(
                'REMOTE_ADDR'
            )
            user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        UserActivity.objects.create(
            user=user,
            action=action,
            module=module,
            description=description,
            ip_address=ip_address,
            user_agent=user_agent,
            metadata=metadata or {},
        )

    @staticmethod
    @transaction.atomic
    def create_user(data: dict, created_by=None) -> User:
        password = data.pop('password', None)
        role_ref = data.pop('role_ref', None)
        role_name = data.pop('role', None)

        if role_ref:
            data['role_ref'] = role_ref
            data['role'] = role_ref.name.upper()
        elif role_name:
            data['role'] = role_name.upper()
            data['role_ref'] = Role.objects.filter(name__iexact=role_name).first()

        if data.get('aadhaar_number'):
            validate_aadhaar(data['aadhaar_number'])
            data['aadhaar_number'] = encrypt_field(data['aadhaar_number'])

        if not data.get('employee_id'):
            data['employee_id'] = generate_employee_id()

        data['created_by'] = created_by
        data['joined_at'] = timezone.now()

        user = User(**data)
        if password:
            user.set_password(password)
        else:
            user.set_unusable_password()
        user.save()
        UserService.log_activity(
            created_by or user,
            'create',
            module='users',
            description=f'Created user {user.email}',
        )
        return user

    @staticmethod
    @transaction.atomic
    def update_user(user: User, data: dict, updated_by=None) -> User:
        password = data.pop('password', None)
        role_ref = data.pop('role_ref', None)
        role_name = data.pop('role', None)

        if role_ref is not None:
            user.role_ref = role_ref
            user.role = role_ref.name.upper()
        elif role_name:
            user.role = role_name.upper()
            user.role_ref = Role.objects.filter(name__iexact=role_name).first()

        if 'aadhaar_number' in data and data['aadhaar_number']:
            validate_aadhaar(data['aadhaar_number'])
            data['aadhaar_number'] = encrypt_field(data['aadhaar_number'])

        for field, value in data.items():
            setattr(user, field, value)

        if password:
            user.set_password(password)

        user.save()
        UserService.log_activity(
            updated_by or user,
            'update',
            module='users',
            description=f'Updated user {user.email}',
        )
        return user

    @staticmethod
    @transaction.atomic
    def deactivate_user(user: User, deactivated_by=None) -> User:
        user.is_active = False
        user.save(update_fields=['is_active', 'updated_at'])
        UserService.log_activity(
            deactivated_by or user,
            'delete',
            module='users',
            description=f'Deactivated user {user.email}',
        )
        return user

    @staticmethod
    def bulk_import_users(file_obj, created_by=None) -> dict:
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        required = {'email', 'phone', 'first_name', 'last_name', 'role'}
        if not required.issubset(set(headers or [])):
            raise ValueError(f'Missing required columns: {required - set(headers or [])}')

        created = 0
        errors: list[dict] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            if not row_data.get('email'):
                continue
            try:
                role = Role.objects.filter(name__iexact=str(row_data.get('role', ''))).first()
                UserService.create_user(
                    {
                        'email': row_data['email'],
                        'phone': str(row_data.get('phone', '')),
                        'first_name': row_data.get('first_name', ''),
                        'last_name': row_data.get('last_name', ''),
                        'role_ref': role,
                        'role': row_data.get('role'),
                        'password': row_data.get('password') or 'ChangeMe@123',
                    },
                    created_by=created_by,
                )
                created += 1
            except Exception as exc:
                errors.append({'row': idx, 'error': str(exc)})
        return {'created': created, 'errors': errors}

    @staticmethod
    def export_users(queryset) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        headers = [
            'employee_id',
            'email',
            'phone',
            'first_name',
            'last_name',
            'role',
            'is_active',
            'college',
        ]
        ws.append(headers)
        for user in queryset.select_related('college', 'role_ref'):
            ws.append(
                [
                    user.employee_id,
                    user.email,
                    user.get_masked_phone(),
                    user.first_name,
                    user.last_name,
                    user.role_ref.display_name if user.role_ref else user.role,
                    user.is_active,
                    user.college.name if user.college else '',
                ]
            )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


class RoleService:
    @staticmethod
    @transaction.atomic
    def create_role(data: dict) -> Role:
        permissions = data.pop('permissions', [])
        role = Role.objects.create(**data)
        RoleService.set_permissions(role, permissions)
        return role

    @staticmethod
    @transaction.atomic
    def update_role(role: Role, data: dict) -> Role:
        permissions = data.pop('permissions', None)
        for field, value in data.items():
            setattr(role, field, value)
        role.save()
        if permissions is not None:
            RoleService.set_permissions(role, permissions)
        return role

    @staticmethod
    def set_permissions(role: Role, permission_ids: list[int]):
        RolePermission.objects.filter(role=role).delete()
        for permission_id in permission_ids:
            RolePermission.objects.create(role=role, module_permission_id=permission_id)

    @staticmethod
    def clone_role(source_role: Role, new_name: str, display_name: str) -> Role:
        clone = Role.objects.create(
            name=new_name,
            display_name=display_name,
            description=f'Cloned from {source_role.display_name}',
            is_system_role=False,
        )
        for rp in source_role.role_permissions.all():
            RolePermission.objects.create(
                role=clone,
                module_permission=rp.module_permission,
                field_restrictions=rp.field_restrictions,
                record_restrictions=rp.record_restrictions,
            )
        return clone

    @staticmethod
    def get_role_permissions_matrix(role: Role) -> list[dict]:
        permissions = ModulePermission.objects.select_related('module').order_by('module__order', 'action')
        assigned = {
            rp.module_permission_id: rp
            for rp in role.role_permissions.select_related('module_permission')
        }
        matrix = []
        for perm in permissions:
            rp = assigned.get(perm.id)
            matrix.append(
                {
                    'id': perm.id,
                    'module': perm.module.name,
                    'module_display': perm.module.display_name,
                    'action': perm.action,
                    'assigned': rp is not None,
                    'field_restrictions': rp.field_restrictions if rp else {},
                    'record_restrictions': rp.record_restrictions if rp else {},
                }
            )
        return matrix
