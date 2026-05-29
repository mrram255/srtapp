from __future__ import annotations

import io
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import Workbook, load_workbook

from apps.core.services import FileUploadService
from apps.core.utils import generate_unique_code
from apps.students.models import Student

User = get_user_model()


class StudentService:
    @staticmethod
    def generate_enrollment_number(department) -> str:
        prefix = getattr(department, 'code', 'STU')[:6].upper()
        return generate_unique_code(prefix, length=6)

    @staticmethod
    def _parse_date(value):
        if value is None or value == '':
            from datetime import date
            return date(2000, 1, 1)
        if hasattr(value, 'date'):
            return value.date() if hasattr(value, 'date') and callable(value.date) else value
        from datetime import date
        text = str(value).strip()[:10]
        try:
            return date.fromisoformat(text)
        except ValueError:
            return date(2000, 1, 1)

    @staticmethod
    @transaction.atomic
    def bulk_import(file_obj, college, created_by=None) -> dict[str, Any]:
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(h).strip().lower() for h in (header_row or []) if h is not None]

        required = {'email', 'phone', 'first_name', 'last_name', 'department_code', 'branch_code'}
        if not required.issubset(set(headers)):
            raise ValueError(f'Missing columns. Required: {", ".join(sorted(required))}')

        created = 0
        errors: list[dict] = []

        from apps.colleges.models import Branch, Department

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            data = dict(zip(headers, row))
            try:
                dept = Department.objects.get(
                    college=college,
                    code=str(data['department_code']).strip(),
                    is_deleted=False,
                )
                branch = Branch.objects.get(
                    college=college,
                    department=dept,
                    code=str(data['branch_code']).strip(),
                    is_deleted=False,
                )
                email = str(data['email']).strip().lower()
                if User.objects.filter(email=email).exists():
                    raise ValueError('Email already exists')

                user = User.objects.create_user(
                    email=email,
                    phone=str(data['phone']).strip(),
                    first_name=str(data['first_name']).strip(),
                    last_name=str(data['last_name']).strip(),
                    role='STUDENT',
                    college=college,
                    password=str(data.get('password') or 'ChangeMe@123'),
                )
                enrollment = str(data.get('enrollment_number') or '').strip()
                if not enrollment:
                    enrollment = StudentService.generate_enrollment_number(dept)

                batch_year = int(data.get('batch_year') or college.established_year or 2024)
                Student.objects.create(
                    user=user,
                    college=college,
                    department=dept,
                    branch=branch,
                    enrollment_number=enrollment,
                    roll_number=str(data.get('roll_number') or enrollment),
                    admission_number=str(data.get('admission_number') or enrollment),
                    semester=int(data.get('semester') or 1),
                    batch_year=batch_year,
                    date_of_birth=StudentService._parse_date(data.get('date_of_birth')),
                    gender=str(data.get('gender') or 'MALE').upper(),
                    address=str(data.get('address') or 'N/A'),
                    city=str(data.get('city') or 'N/A'),
                    state=str(data.get('state') or 'N/A'),
                    pincode=str(data.get('pincode') or '000000'),
                    emergency_contact=str(data.get('emergency_contact') or '0000000000'),
                    emergency_contact_name=str(data.get('emergency_contact_name') or 'Guardian'),
                    admission_date=StudentService._parse_date(data.get('admission_date') or '2024-07-01'),
                    category=str(data.get('category') or ''),
                    abc_id=str(data.get('abc_id') or ''),
                    family_details={},
                    education_details={},
                )
                created += 1
            except Exception as exc:
                errors.append({'row': idx, 'error': str(exc)})

        return {'created': created, 'errors': errors}

    @staticmethod
    def export_students(queryset) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Students'
        ws.append(
            [
                'enrollment_number',
                'roll_number',
                'full_name',
                'email',
                'phone',
                'department',
                'branch',
                'semester',
                'category',
                'student_status',
                'abc_id',
            ]
        )
        for student in queryset.select_related('user', 'department', 'branch'):
            ws.append(
                [
                    student.enrollment_number,
                    student.roll_number,
                    student.user.get_full_name(),
                    student.user.email,
                    student.user.phone,
                    student.department.name,
                    student.branch.name,
                    student.semester,
                    student.category,
                    student.student_status,
                    student.abc_id,
                ]
            )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def import_template() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                'email',
                'phone',
                'first_name',
                'last_name',
                'password',
                'department_code',
                'branch_code',
                'enrollment_number',
                'roll_number',
                'semester',
                'batch_year',
                'date_of_birth',
                'gender',
                'address',
                'city',
                'state',
                'pincode',
                'emergency_contact',
                'emergency_contact_name',
                'admission_date',
                'category',
            ]
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def upload_student_file(file_obj, folder: str) -> str:
        return FileUploadService.upload(file_obj, folder=folder)
